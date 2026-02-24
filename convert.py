"""엑셀 시간표 데이터를 data.js로 변환하는 스크립트"""
import openpyxl
import json

wb = openpyxl.load_workbook('전체교사시간표 최종(웹소스파일) 과목명 풀.xlsx', data_only=True)
ws = wb.active

# 열 매핑 (헤더에서 확인한 구조)
SCHEDULE = {
    "월": list(range(3, 10)),    # C~I (7교시)
    "화": list(range(10, 17)),   # J~P (7교시)
    "수": list(range(17, 23)),   # Q~V (6교시)
    "목": list(range(23, 30)),   # W~AC (7교시)
    "금": list(range(30, 37)),   # AD~AJ (7교시)
}

teachers_data = {}
classes_data = {}
teacher_list = []

# 교사 데이터 파싱 (2행 단위: 과목행 + 학반행)
row = 3  # 데이터 시작 (1,2행은 헤더)
while row <= ws.max_row:
    teacher_num = ws.cell(row=row, column=1).value
    teacher_name = ws.cell(row=row, column=2).value

    if teacher_num is None or teacher_name is None:
        row += 1
        continue

    teacher_name = str(teacher_name).strip()
    teacher_list.append(teacher_name)

    subject_row = row
    class_row = row + 1

    teacher_schedule = {}

    for day, cols in SCHEDULE.items():
        day_schedule = {}
        for period_idx, col in enumerate(cols):
            period = period_idx + 1
            subject = ws.cell(row=subject_row, column=col).value
            class_name = ws.cell(row=class_row, column=col).value

            if subject and class_name:
                subject = str(subject).strip()
                class_name = str(class_name).strip()
                day_schedule[period] = {"subject": subject, "class": class_name}

                # 학반 시간표에도 추가
                if class_name not in classes_data:
                    classes_data[class_name] = {}
                if day not in classes_data[class_name]:
                    classes_data[class_name][day] = {}

                entry = {"subject": subject, "teacher": teacher_name}
                existing = classes_data[class_name][day].get(period)
                if existing:
                    if isinstance(existing, list):
                        existing.append(entry)
                    else:
                        classes_data[class_name][day][period] = [existing, entry]
                else:
                    classes_data[class_name][day][period] = entry
            else:
                day_schedule[period] = None

        teacher_schedule[day] = day_schedule

    teachers_data[teacher_name] = teacher_schedule
    row += 2  # 다음 교사 (2행 단위)

# 학반 목록 정렬
import re
def class_sort_key(c):
    parts = c.split('-')
    return (int(parts[0]), int(parts[1]))

class_list = sorted([c for c in classes_data if re.match(r'^\d+-\d+$', c)], key=class_sort_key)
teacher_list.sort()

output = {
    "teachers": teachers_data,
    "classes": classes_data,
    "teacherList": teacher_list,
    "classList": class_list,
}

with open('data.js', 'w', encoding='utf-8') as f:
    f.write('const TIMETABLE_DATA = ')
    json.dump(output, f, ensure_ascii=False, indent=None)
    f.write(';\n')

print(f"변환 완료: 교사 {len(teacher_list)}명, 학반 {len(class_list)}개")
print(f"교사 목록: {teacher_list[:10]}")
print(f"학반 목록: {class_list}")
